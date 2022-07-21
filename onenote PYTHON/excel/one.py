import openpyxl

path = 'excel.xlsx'
workbook = openpyxl.load_workbook(path)

sheet = workbook.active
sheet_di = sheet.dimensions
print(f"当前活动的表格是{sheet.title}")
cell1 = sheet['A1:B2']

ex_max_row = sheet.max_row
ex_max = sheet.max_column

for i in sheet.iter_rows():
    for j in i:
        print(j.value)
print(ex_max , ex_max_row)

sheet['A1'] = ''

workbook.save('test.xlsx')